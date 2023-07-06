from pathlib import Path
from openpyxl import Workbook, load_workbook
from tkinter import Tk, Canvas, Button

OUTPUT_PATH = Path(__file__).parent

projects_file_path = "projects.xlsx"

def create_excel_file():
    '''Check if the projects file already exists, and if not, create a new Excel workbook,
    save it as the projects file, create the "Employees" sheet, and print a message confirming the creation.'''
    if not Path(projects_file_path).is_file():
        workbook = Workbook()
        workbook.save(projects_file_path)
        workbook.create_sheet(title="Employees")  # Create the "Employees" sheet
        workbook.save(projects_file_path)
        print("Projects file created with 'Employees' sheet.")

def create_project():
    '''Prompt the user to enter a project name, check if the project already exists in the workbook,
 and if not, create a new sheet with the project name, save the workbook, 
 and print a success message.'''
    #project_name = input("Enter the project name: ")
    workbook = load_workbook(projects_file_path)
    if project_name in workbook.sheetnames:
        flag = False
        #print("Project already exists.")
        return flag

    workbook.create_sheet(title=project_name)
    workbook.save(projects_file_path)
    flag = True
    return flag
    #print("New project created successfully!")

def assign_employee(project_name, employee_id):
    '''Load the workbook, check if the specified project exists, if yes, append the employee ID 
    to the tasks sheet of the project, save the workbook, and print a success message.
    If the project does not exist, print an error message.'''
    workbook = load_workbook(projects_file_path)
    if project_name in workbook.sheetnames:
        tasks_sheet = workbook[project_name]
        tasks_sheet.append([f"ID: {employee_id}"])
        workbook.save(projects_file_path)
        print("Employee assigned to project successfully!")
    else:
        print("Project not found.")

def add_task(project_name, employee_id, task, priority):
    workbook = load_workbook(projects_file_path)
    if project_name in workbook.sheetnames:
        tasks_sheet = workbook[project_name]
        tasks_sheet.append([f"ID: {employee_id}", task, priority, "Incomplete"])
        workbook.save(projects_file_path)
        print("Task added successfully!")
    else:
        print("Project not found.")

def mark_complete(project_name, employee_id, task):
    ''' Load the workbook, check if the specified project exists, if yes, append the task 
    details (employee ID, task, priority, and status) to the tasks sheet of the project, 
    save the workbook, and print a success message. 
    If the project does not exist, print an error message.'''
    workbook = load_workbook(projects_file_path)
    if project_name in workbook.sheetnames:
        tasks_sheet = workbook[project_name]
        for row in tasks_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == f"ID: {employee_id}" and row[1] == task:
                tasks_sheet.cell(row=row[0].row, column=4).value = "Complete"
                workbook.save(projects_file_path)
                print("Task marked as complete.")
                return
        print("Task not found.")
    else:
        print("Project not found.")

def display_tasks(project_name, employee_id):
    ''' Load the workbook, check if the specified project exists, if yes, iterate through 
    the tasks sheet of the project to find tasks assigned to the specified employee ID. 
    Append the task details (task and status) to the tasks list. 
    If tasks are found, print them with a header. If no tasks are found,
    print a corresponding message. If the project does not exist, print an error message.'''
    workbook = load_workbook(projects_file_path)
    if project_name in workbook.sheetnames:
        tasks_sheet = workbook[project_name]
        tasks = []
        for row in tasks_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == f"ID: {employee_id}":
                task = row[1]
                status = row[3]
                tasks.append(f"Task: {task}, Status: {status}")
        if tasks:
            print(f"Tasks for Employee ID {employee_id}:")
            print("\n".join(tasks))
        else:
            print("No tasks found for the employee.")
    else:
        print("Project not found.")

def create_employee(name, surname, dob):
    '''Load the workbook, create an "Employees" sheet if it doesn't exist, retrieve the next
    available employee ID, append the employee details (ID, name, surname, and date of birth)
    to the "Employees" sheet, save the workbook, and print a success message.'''
    if len(dob) != 10:
        flag = False
        #print("Project already exists.")
        return flag
    
    workbook = load_workbook(projects_file_path)
    employee_sheet = workbook.create_sheet(title="Employees")
    next_employee_id = get_next_employee_id()
    employee_sheet.append([f"ID: {next_employee_id}", name, surname, dob])
    workbook.save(projects_file_path)
    #gui section
    flag = True
    return flag

def get_next_employee_id():
    '''Load the workbook, retrieve the "Employees" sheet, iterate over the rows starting
    from the second row, extract the employee ID from each row, find the maximum ID, 
    and return the next available employee ID by incrementing the maximum ID by 1.'''
    workbook = load_workbook(projects_file_path)
    employee_sheet = workbook["Employees"]
    max_id = 0
    for row in employee_sheet.iter_rows(min_row=2, values_only=True):
        employee_id = row[0].split(":")[1].strip()
        if int(employee_id) > max_id:
            max_id = int(employee_id)
    return max_id + 1

def get_all_projects():
    '''# Load the workbook, retrieve the names of all the sheets (projects), 
    remove the "Employees" sheet, and return the list of projects.'''
    workbook = load_workbook(projects_file_path)
    projects = workbook.sheetnames
    projects.remove("Employees")
    return projects

def get_all_employees():
    '''Load the workbook, retrieve the "Employees" sheet, iterate through its rows to 
    collect employee information, and return a list of formatted employee details.'''
    workbook = load_workbook(projects_file_path)
    employee_sheet = workbook["Employees"]
    employees = []
    for row in employee_sheet.iter_rows(min_row=2, values_only=True):
        employee_id = row[0]
        name = row[1]
        surname = row[2]
        employees.append(f"ID: {employee_id}, Name: {name} {surname}")
    return employees

def overview_projects():
    '''Retrieve all projects, iterate through each project, and print the project name
    along with the associated tasks and their statuses.'''
    projects = get_all_projects()
    if len(projects) > 0:
        print("Projects Overview:")
        for project in projects:
            print(f"Project: {project}")
    projects = get_all_projects()
    if len(projects) > 0:
        print("Projects Overview:")
        for project in projects:
            print(f"Project: {project}")
            tasks_sheet = workbook[project]
            if tasks_sheet is not None:
                for row in tasks_sheet.iter_rows(min_row=2, values_only=True):
                    employee_id = row[0]
                    task = row[1]
                    status = row[3]
                    print(f"  - Employee ID: {employee_id}, Task: {task}, Status: {status}")
            else:
                print("  - No tasks found for this project.")
            print()
    else:
        print("No projects found.")
        
def overview_employees():
    '''Retrieve the employee sheet from the workbook, iterate through each employee, 
    and print their ID, name, and surname if employees exist; otherwise, 
    display appropriate messages.'''
    workbook = load_workbook(projects_file_path)
    if "Employees" in workbook.sheetnames:
        employee_sheet = workbook["Employees"]
        employees = []
        for row in employee_sheet.iter_rows(min_row=2, values_only=True):
            employee_id = row[0]
            name = row[1]
            surname = row[2]
            employees.append(f"ID: {employee_id}, Name: {name} {surname}")
        if employees:
            print("Employees Overview:")
            print("\n".join(employees))
        else:
            print("No employees found.")
    else:
        print("Employees sheet not found.")

def overview_tasks():
    ''' Retrieve the tasks for each project in the workbook and display an overview of 
    tasks including the project name, employee ID, task, and status;
    if no tasks exist for a project, display appropriate messages.'''
    workbook = load_workbook(projects_file_path)
    projects = workbook.sheetnames
    if "Employees" in projects:
        projects.remove("Employees")
    workbook = load_workbook(projects_file_path)
    projects = workbook.sheetnames
    if "Employees" in projects:
        projects.remove("Employees")

    if len(projects) > 0:
        print("Task Overview:")
        for project in projects:
            tasks_sheet = workbook[project]
            if tasks_sheet is not None:
                for row in tasks_sheet.iter_rows(min_row=2, values_only=True):
                    employee_id = row[0]
                    task = row[1]
                    status = row[3]
                    print(f"Project: {project}, Employee ID: {employee_id}")
                    print(f"  - Task: {task}, Status: {status}")
                    print()
            else:
                print(f"No tasks found for project: {project}")
                print()
    else:
        print("No tasks found for any employee.")

def list_proj():
    wb = load_workbook("projects.xlsx")
    proj_list = list()
    for sheet in wb.worksheets:
        proj_list.append(sheet.title)
    return proj_list

def main():
    '''Main function that provides a menu-driven interface for project management and overview,
    allowing users to create projects, assign employees, add tasks, mark tasks as complete, 
    display tasks for employees, create employees, and view project, employee, 
    and task overviews.'''
    create_excel_file()
    while True:
        print("\n=== PROJECT MANAGEMENT ===")
        print("1. Create a new project")
        print("2. Assign an employee to a project")
        print("3. Add a task for an employee")
        print("4. Mark a task as complete")
        print("5. Display tasks for an employee")
        print("6. Create a new employee")
        print("\n=== OVERVIEW SECTION ===")
        print("7. Overview of all projects")
        print("8. Overview of all employees")
        print("9. Overview of tasks for each employee")
        print("\n==========================")
        print("10. Exit")
        print("11 List Task")

        choice = input("Enter your choice (1-10): ")

        if choice == "1":
            create_project()
        elif choice == "2":
            project = input("Enter the project name: ")
            employee_id = input("Enter the employee ID: ")
            assign_employee(project, employee_id)
        elif choice == "3":
            project = input("Enter the project name: ")
            employee_id = input("Enter the employee ID: ")
            task = input("Enter the task: ")
            priority = input("Enter the priority: ")
            add_task(project, employee_id, task, priority)
        elif choice == "4":
            project = input("Enter the project name: ")
            employee_id = input("Enter the employee ID: ")
            task = input("Enter the task: ")
            mark_complete(project, employee_id, task)
        elif choice == "5":
            project = input("Enter the project name: ")
            employee_id = input("Enter the employee ID: ")
            display_tasks(project, employee_id)
        elif choice == "6":
            name = input("Enter the employee name: ")
            surname = input("Enter the employee surname: ")
            dob = input("Enter the employee date of birth: ")
            create_employee(name, surname, dob)
        elif choice == "7":
            overview_projects()
        elif choice == "8":
            overview_employees()
        elif choice == "9":
            overview_tasks()
        elif choice == "10":
            break
        elif choice == "11":
            list_task()
        else:
            print("Invalid choice. Please try again.")
            

if __name__ == "__main__":
    main()