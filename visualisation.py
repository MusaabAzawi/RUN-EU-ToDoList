from pathlib import Path
from openpyxl import Workbook, load_workbook
from tkinter import Tk, Canvas, Button

OUTPUT_PATH = Path(__file__).parent

projects_file_path = "projects.xlsx"
employees_file_path = "employees.xlsx"

def create_project():
    project_name = input("Enter the project name: ")
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet(title=project_name)
    workbook.save(projects_file_path)
    print("New project created successfully!")

def get_all_projects():
    workbook = load_workbook(projects_file_path)
    projects = workbook.sheet_names
    if projects:
        print("List of projects:")
        print("\n".join(projects))
    else:
        print("No projects found.")

def main():
    # Create necessary files if they don't exist
    if not Path(projects_file_path).exists():
        workbook = Workbook()
        workbook.save(projects_file_path)

    if not Path(employees_file_path).exists():
        workbook = Workbook()
        workbook.create_sheet(title="Employees")
        workbook.active.append(["ID", "Name", "Surname", "DOB"])
        workbook.save(employees_file_path)

    window = Tk()
    window.geometry("1280x720")
    window.configure(bg="#FFFFFF")

    canvas = Canvas(
        window,
        bg="#FFFFFF",
        height=720,
        width=1280,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )
    canvas.place(x=0, y=0)

    button_1 = Button(
        text="New Project",
        command=create_project
    )
    button_1.place(
        x=73.0,
        y=582.0,
        width=390.0,
        height=57.0
    )

    canvas.create_text(
        62.0,
        46.0,
        anchor="nw",
        text="List Manager",
        fill="#000000",
        font=("Helvetica Bold", 64 * -1)
    )

    canvas.create_text(
        62.0,
        123.0,
        anchor="nw",
        text="A RUN-EU Project",
        fill="#000000",
        font=("Helvetica Light", 32 * -1)
    )

    canvas.create_rectangle(
        559.0,
        46.0,
        1234.0,
        674.0,
        fill="#D9D9D9",
        outline=""
    )

    button_2 = Button(
        text="Get All Projects",
        command=get_all_projects
    )
    button_2.place(
        x=73.0,
        y=493.0,
        width=390.0,
        height=57.0
    )

    window.resizable(False, False)
    window.mainloop()

if __name__ == "__main__":
    main()